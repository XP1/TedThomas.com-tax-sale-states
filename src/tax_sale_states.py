#!/usr/bin/env python
"""
Copyright 2024 XP1

Licensed under the Apache License, Version 2.0 (the "License");
you may not use this file except in compliance with the License.
You may obtain a copy of the License at

    http://www.apache.org/licenses/LICENSE-2.0

Unless required by applicable law or agreed to in writing, software
distributed under the License is distributed on an "AS IS" BASIS,
WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
See the License for the specific language governing permissions and
limitations under the License.
"""

from contextlib import closing
from zipfile import ZipFile, ZIP_DEFLATED
from io import BytesIO
import abc
import csv
import json
import os
import time
import datetime
import itertools

from bs4 import BeautifulSoup
from py_markdown_table.markdown_table import markdown_table

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles.numbers import FORMAT_TEXT
from openpyxl.worksheet.table import Table, TableColumn, TableStyleInfo
from openpyxl.worksheet.filters import AutoFilter
from openpyxl.writer import theme, excel

import requests
from urllib3.exceptions import InsecureRequestWarning

requests.packages.urllib3.disable_warnings(category=InsecureRequestWarning)

session = requests.Session()
session.verify = False
session.headers.update({
    "accept-encoding": "gzip, deflate, br"
})


def fetch_text_file(path: str) -> str:
    with open(path, encoding="utf-8") as file:
        text = file.read()
        return text


def write_text_file(path: str, data: any) -> None:
    with open(path, "w", encoding="utf-8") as file:
        file.write(data)


def write_json_file(path: str, data: any) -> None:
    with open(path, "w", encoding="utf8") as file:
        json.dump(data, file, indent=4)


class WorkbookController():
    XL_FOLDER_NAME = "xl"

    CONTENT_TYPES_XML_FILE_NAME = "[Content_Types].xml"
    WORKBOOK_XML_FILE_NAME = "workbook.xml"
    STYLES_XML_FILE_NAME = "styles.xml"

    FIRST_NAMES = [
        CONTENT_TYPES_XML_FILE_NAME,
        f"{XL_FOLDER_NAME}/{WORKBOOK_XML_FILE_NAME}",
        f"{XL_FOLDER_NAME}/{STYLES_XML_FILE_NAME}"
    ]

    ISO_8601_NUMBER_FORMAT = "yyyy-mm-ddThh:MM:ss"

    was_theme_updated = False

    def __init__(self) -> None:
        cls = self.__class__
        if not cls.was_theme_updated:
            self.update_theme()
            cls.was_theme_updated = True

    @staticmethod
    def update_theme() -> None:
        color = {
            "1F497D": "44546A",
            "EEECE1": "E7E6E6",
            "4F81BD": "5B9BD5",
            "C0504D": "ED7D31",
            "9BBB59": "A5A5A5",
            "8064A2": "FFC000",
            "4BACC6": "4472C4",
            "F79646": "70AD47",
            "0000FF": "0563C1",
            "800080": "954F72"
        }

        xml = theme.theme_xml
        for original, replacement in color.items():
            xml = xml.replace(f"val=\"{original}\"", f"val=\"{replacement}\"")

        theme.theme_xml = xml
        excel.theme_xml = xml

    @classmethod
    def fix_workbook_mime_type(cls, file_path):
        buffer = BytesIO()

        with ZipFile(file_path) as zip_file:
            names = zip_file.namelist()

            FIRST_NAMES = cls.FIRST_NAMES
            remaining_names = [name for name in names if name not in FIRST_NAMES]
            ordered_names = FIRST_NAMES + remaining_names

            with ZipFile(buffer, "w", ZIP_DEFLATED, allowZip64=True) as buffer_zip_file:
                for name in ordered_names:
                    try:
                        file = zip_file.open(name)
                        buffer_zip_file.writestr(file.name, file.read())
                    except KeyError:
                        pass

        return buffer

    @staticmethod
    def autosize_columns(worksheet):
        def value_of(value):
            return (str(value) if value is not None else "")

        for cells in worksheet.columns:
            length = max(len(value_of(cell.value)) for cell in cells)
            column_letter = get_column_letter(cells[0].column)
            worksheet.column_dimensions[column_letter].width = length

        return worksheet

    @classmethod
    def create_workbook(
        cls,

        key_title: dict = {},
        data: list = [],

        workbook_title: str = "Workbook",
        sheet_title: str = "Sheet",
        table_name: str = "Table"
    ):
        keys = key_title.keys()
        titles = list(key_title.values())

        # Initialize workbook and worksheet.
        workbook = Workbook()
        properties = workbook.properties
        properties.title = workbook_title
        properties.creator = None

        sheet = workbook.active
        sheet.title = sheet_title

        # Add data.
        rows = [key_title] + data
        for r, row in enumerate(rows, start=1):
            for c, key in enumerate(keys, start=1):
                value = (row[key] if key in row else "")
                cell = sheet.cell(row=r, column=c)

                cell_number_format = FORMAT_TEXT
                cell_value = value
                cell_data_type = "s"
                if isinstance(value, datetime.datetime):
                    cell_number_format = cls.ISO_8601_NUMBER_FORMAT
                    cell_data_type = "d"

                cell.number_format = cell_number_format
                cell.value = cell_value
                cell.data_type = cell_data_type

        total_row = [""] * len(keys)
        total_row[0] = "Total"
        total_row[-1] = f"=SUBTOTAL(103,{table_name}[{titles[-1]}])"
        sheet.append(total_row)

        # Add table.
        style = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )

        table_columns = tuple(TableColumn(id=h, name=header) for h, header in enumerate(titles, start=1))
        total_column = table_columns[0]
        total_column.totalsRowLabel = "Total"
        count_column = table_columns[-1]
        count_column.totalsRowFunction = "count"

        max_column = sheet.max_column
        max_column_letter = get_column_letter(max_column)
        max_row = sheet.max_row
        table = Table(
            displayName=table_name,
            ref=f"A1:{max_column_letter}{max_row}",
            autoFilter=AutoFilter(ref=f"A1:{max_column_letter}{max_row - 1}"),
            tableStyleInfo=style,
            totalsRowShown=True,
            totalsRowCount=1,
            tableColumns=table_columns
        )

        sheet.add_table(table)

        # Adjust column sizes.
        cls.autosize_columns(sheet)

        # Set the active cell under the table.
        active_cell = f"A{max_row + 1}"
        selection = sheet.sheet_view.selection[0]
        selection.activeCell = active_cell
        selection.sqref = active_cell

        return workbook


class TaxSaleStates(metaclass=abc.ABCMeta):
    def __init__(self, name: str, uri: str, key_title: dict) -> None:
        super().__init__()

        self.name = name
        self.uri = uri

        data_path = f"../data"
        self.data_path = data_path

        self.html_file_path = f"{data_path}/{name}.html"

        build_path = f"../build/{name}"
        self.build_path = build_path

        self.excel_file_path = f"{build_path}/{name}.xlsx"
        self.csv_file_path = f"{build_path}/{name}.csv"
        self.json_file_path = f"{build_path}/{name}.json"
        self.markdown_file_path = f"{build_path}/{name}.md"

        self.key_title = key_title

        titles = key_title.values()
        self.titles = titles

        self.title_title = {title: title for title in titles}

        self.markup = None

        self.data = None

    def add_missing_keys(self, row: dict) -> dict:
        titles = self.titles
        for title in titles:
            if title not in row:
                row[title] = ""
        return row

    def parse_table(self, table) -> dict:
        key_title = self.key_title

        rows = table.select("tr")

        result = {}

        for row in rows:
            cells = row.select("td")
            if len(cells) < 2:
                continue

            key = cells[0].get_text().strip().rstrip(":")
            value = cells[1].get_text().strip()
            if len(key) <= 0 and len(value) <= 0:
                continue

            title = (key_title[key] if key in key_title else key)
            result[title] = value

        return result

    def parse(self) -> list:
        items = []

        soup = BeautifulSoup(self.markup, "html.parser")
        # State headings.
        state_headings = soup.select(".elementor-widget-menu-anchor + .elementor-widget-heading .elementor-heading-title")
        # states = [state_heading.get_text().strip() for state_heading in state_headings]

        # Table and notes.
        table_notes = iter(soup.select(".e-con-inner > .e-child > .e-child .elementor-widget-text-editor"))
        i = 0
        for table, notes in itertools.zip_longest(table_notes, table_notes):
            state = state_headings[i].get_text().strip()
            table_items = self.parse_table(table)
            description = notes.get_text().strip().removeprefix("NOTES:").lstrip()

            item = self.add_missing_keys({
                "State": state,
                **table_items,
                "Description": description
            })
            items.append(item)
            i += 1

        self.data = items

        return items


KEY_TITLE = {
    "State": "State",
    "Type": "Type",
    "Bidding Process": "Bidding process",
    "Frequency": "Frequency",
    "Interest Rate / Penalty": "Interest rate / penalty",
    "Redemption Period": "Redemption period",
    "Online Auction": "Online auction",
    "Over the Counter": "Over the counter",
    "Statute": "Statute",
    "Notes": "Notes",
    "Description": "Description"
}


class TaxLienCertificateStates(TaxSaleStates):
    def __init__(self) -> None:
        name = "Tax lien certificate states"
        uri = "https://tedthomas.com/faqs/tax-lien-certificate-states/"
        super().__init__(name, uri, KEY_TITLE)


class TaxDeedStates(TaxSaleStates):
    def __init__(self) -> None:
        name = "Tax deed states"
        uri = "https://tedthomas.com/faqs/tax-deed-states/"
        super().__init__(name, uri, KEY_TITLE)


class MainController():
    def __init__(self) -> None:
        super().__init__()

        self.workbook_buffer = None

    def run(self) -> None:
        models = [TaxLienCertificateStates(), TaxDeedStates()]
        for model in models:
            self.build(model)
            print()

    def build(self, model) -> None:
        print(f"Building \"{model.name}\"...")

        print("    Fetching data...", end="", flush=True)
        model.markup = self.fetch_markup(model)
        print(" Done.")

        os.makedirs(model.data_path, exist_ok=True)

        print("    Writing data...", end="", flush=True)
        self.write_markup(model)
        print(" Done.")

        model.parse()

        print("    Creating workbook...", end="", flush=True)
        workbook_buffer = self.create_workbook(model)
        print(" Done.")

        os.makedirs(model.build_path, exist_ok=True)

        print("    Writing workbook...", end="", flush=True)
        self.write_workbook(model, workbook_buffer)
        print(" Done.")

        print("    Writing CSV...", end="", flush=True)
        self.write_csv(model)
        print(" Done.")

        print("    Writing JSON...", end="", flush=True)
        self.write_json(model)
        print(" Done.")

        print("    Writing markdown...", end="", flush=True)
        self.write_markdown(model)
        print(" Done.")

    def fetch_markup(self, model) -> str:
        response = session.get(url=model.uri, timeout=5)

        status_code = response.status_code
        if not (status_code == requests.codes.ok or status_code == requests.codes.partial_content):
            print(f"    Received status code {status_code} while fetching {model.name} data. Retrying...")
            time.sleep(5)
            return self.fetch_markup(model)

        return response.text

    @staticmethod
    def write_markup(model) -> None:
        write_text_file(model.html_file_path, model.markup)

    def create_workbook(self, model):
        name = model.name
        workbook_controller = WorkbookController()
        with closing(workbook_controller.create_workbook(
            key_title=model.title_title,
            data=model.data,

            workbook_title=name,
            sheet_title=name,
            table_name=type(model).__name__
        )) as workbook:
            workbook_buffer = BytesIO()
            workbook.save(workbook_buffer)
            workbook_buffer = workbook_controller.fix_workbook_mime_type(workbook_buffer)
            self.workbook_buffer = workbook_buffer
            return workbook_buffer

    def write_workbook(self, model, workbook_buffer=None) -> None:
        with open(model.excel_file_path, "wb") as file:
            if workbook_buffer is None:
                workbook_buffer = self.workbook_buffer

            file.write(workbook_buffer.getvalue())

    @staticmethod
    def write_csv(model) -> None:
        with open(model.csv_file_path, "w") as file:
            writer = csv.DictWriter(file, fieldnames=model.titles)
            writer.writeheader()
            writer.writerows(model.data)

    @staticmethod
    def write_json(model) -> None:
        write_json_file(model.json_file_path, model.data)

    @staticmethod
    def write_markdown(model) -> None:
        markdown = markdown_table(model.data).get_markdown()
        write_text_file(model.markdown_file_path, markdown)


def main(argv: list[str]) -> None:
    main_controller = MainController()
    main_controller.run()


if __name__ == "__main__":
    main([])